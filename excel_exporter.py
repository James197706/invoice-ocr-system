"""
發票 Excel 匯出模組
Invoice Excel Exporter

產生包含三個工作表的專業 Excel 報表：
  1. 發票彙總   - 所有發票的摘要列表
  2. 品項明細   - 各發票的品項逐筆明細
  3. 統計分析   - 依科目/部門/廠商的統計
"""

from __future__ import annotations

import io
from datetime import datetime
from typing import Any

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    GradientFill,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference


# ─────────────────────────────────────────
# 顏色常數
# ─────────────────────────────────────────
CLR_HEADER_BG   = "1E3A8A"   # 深藍（標題列背景）
CLR_HEADER_FG   = "FFFFFF"   # 白色（標題列文字）
CLR_SUBHDR_BG   = "DBEAFE"   # 淡藍（子標題列背景）
CLR_SUBHDR_FG   = "1E40AF"   # 藍色（子標題列文字）
CLR_ALT_ROW     = "F0F7FF"   # 隔行底色
CLR_TOTAL_BG    = "FEF3C7"   # 合計列底色（淡黃）
CLR_TOTAL_FG    = "92400E"   # 合計列文字（棕）
CLR_BORDER      = "CBD5E1"   # 框線色
CLR_ACCENT      = "10B981"   # 綠色強調
CLR_WARNING_BG  = "FFF3CD"   # 警告底色


def _thin_border() -> Border:
    side = Side(style="thin", color=CLR_BORDER)
    return Border(left=side, right=side, top=side, bottom=side)


def _thick_bottom() -> Border:
    thin = Side(style="thin", color=CLR_BORDER)
    thick = Side(style="medium", color="94A3B8")
    return Border(left=thin, right=thin, top=thin, bottom=thick)


def _header_font(size: int = 10) -> Font:
    return Font(name="Arial", bold=True, color=CLR_HEADER_FG, size=size)


def _body_font(size: int = 10, bold: bool = False, color: str = "1F2937") -> Font:
    return Font(name="Arial", size=size, bold=bold, color=color)


def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def _left(wrap: bool = False) -> Alignment:
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)


def _right() -> Alignment:
    return Alignment(horizontal="right", vertical="center")


def _fill(color: str) -> PatternFill:
    return PatternFill("solid", start_color=color, end_color=color)


def _set_row_height(ws, row: int, height: float):
    ws.row_dimensions[row].height = height


def _safe_float(value) -> float:
    """安全轉換為浮點數"""
    try:
        return float(str(value).replace(",", ""))
    except (ValueError, TypeError):
        return 0.0


# ─────────────────────────────────────────
# 工作表 1：發票彙總
# ─────────────────────────────────────────

SUMMARY_COLS = [
    ("序號",        "center", 5),
    ("狀態",        "center", 8),
    ("發票類型",    "center", 14),
    ("發票號碼",    "center", 16),
    ("發票日期",    "center", 13),
    ("賣方名稱",    "left",   22),
    ("賣方統編",    "center", 12),
    ("買方名稱",    "left",   22),
    ("買方統編",    "center", 12),
    ("費用科目",    "left",   14),
    ("部門",        "left",   10),
    ("專案代碼",    "center", 12),
    ("幣別",        "center", 6),
    ("未稅金額",    "right",  13),
    ("稅率(%)",     "center", 9),
    ("稅額",        "right",  11),
    ("含稅總金額",  "right",  14),
    ("付款方式",    "center", 10),
    ("備註",        "left",   30),
    ("來源檔案",    "left",   20),
    ("辨識時間",    "center", 18),
]


def _build_summary_sheet(ws, invoices: list[dict], title: str,
                          company: str, period: str, preparer: str):
    # ── 報表標頭 ──────────────────────────────────
    ws.merge_cells("A1:U1")
    ws["A1"] = f"🧾 {title}"
    ws["A1"].font = Font(name="Arial", bold=True, size=16, color=CLR_HEADER_FG)
    ws["A1"].fill = _fill(CLR_HEADER_BG)
    ws["A1"].alignment = _center()
    _set_row_height(ws, 1, 36)

    ws.merge_cells("A2:U2")
    ws["A2"] = (
        f"公司：{company}　　"
        f"期間：{period}　　"
        f"製表人：{preparer}　　"
        f"製表日期：{datetime.now().strftime('%Y/%m/%d %H:%M')}"
    )
    ws["A2"].font = Font(name="Arial", size=10, italic=True, color="64748B")
    ws["A2"].alignment = _center()
    ws["A2"].fill = _fill(CLR_SUBHDR_BG)
    _set_row_height(ws, 2, 22)

    # ── 欄位標題 ──────────────────────────────────
    HEADER_ROW = 3
    for col_idx, (label, align, width) in enumerate(SUMMARY_COLS, start=1):
        cell = ws.cell(row=HEADER_ROW, column=col_idx, value=label)
        cell.font = Font(name="Arial", bold=True, color=CLR_HEADER_FG, size=10)
        cell.fill = _fill("2563EB")
        cell.alignment = _center()
        cell.border = _thin_border()
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    _set_row_height(ws, HEADER_ROW, 24)

    # ── 資料列 ──────────────────────────────────
    DATA_START = HEADER_ROW + 1
    total_twn = 0.0

    for row_offset, inv in enumerate(invoices):
        r = DATA_START + row_offset
        bg = CLR_ALT_ROW if row_offset % 2 == 0 else "FFFFFF"
        row_fill = _fill(bg)

        total_val = _safe_float(inv.get("total_amount", 0))
        total_twn += total_val if inv.get("currency", "TWD") == "TWD" else 0

        row_data = [
            row_offset + 1,
            inv.get("status", ""),
            inv.get("invoice_type", ""),
            inv.get("invoice_number", ""),
            inv.get("invoice_date", ""),
            inv.get("seller_name", ""),
            inv.get("seller_tax_id", ""),
            inv.get("buyer_name", ""),
            inv.get("buyer_tax_id", ""),
            inv.get("account_category", ""),
            inv.get("department", ""),
            inv.get("project_code", ""),
            inv.get("currency", "TWD"),
            _safe_float(inv.get("subtotal", 0)) or "",
            inv.get("tax_rate", "5"),
            _safe_float(inv.get("tax_amount", 0)) or "",
            _safe_float(inv.get("total_amount", 0)) or "",
            inv.get("payment_method", ""),
            inv.get("notes", ""),
            inv.get("source_file", ""),
            inv.get("recognized_at", ""),
        ]

        aligns = [a for _, a, _ in SUMMARY_COLS]

        for col_idx, (val, align) in enumerate(zip(row_data, aligns), start=1):
            cell = ws.cell(row=r, column=col_idx, value=val)
            cell.font = _body_font()
            cell.fill = row_fill
            cell.border = _thin_border()
            if align == "center":
                cell.alignment = _center()
            elif align == "right":
                cell.alignment = _right()
            else:
                cell.alignment = _left(wrap=True)

            # 金額格式
            if col_idx in (14, 16, 17) and isinstance(val, float) and val:
                cell.number_format = '#,##0.00'

        _set_row_height(ws, r, 18)

    # ── 合計列 ──────────────────────────────────
    total_row = DATA_START + len(invoices)
    ws.merge_cells(f"A{total_row}:M{total_row}")
    total_cell = ws[f"A{total_row}"]
    total_cell.value = f"合計（共 {len(invoices)} 筆）"
    total_cell.font = Font(name="Arial", bold=True, size=11, color=CLR_TOTAL_FG)
    total_cell.fill = _fill(CLR_TOTAL_BG)
    total_cell.alignment = _center()
    total_cell.border = _thin_border()

    # 未稅合計
    subtotal_sum = sum(_safe_float(inv.get("subtotal", 0)) for inv in invoices)
    tax_sum = sum(_safe_float(inv.get("tax_amount", 0)) for inv in invoices)
    total_sum = sum(_safe_float(inv.get("total_amount", 0)) for inv in invoices)

    for col, val in [(14, subtotal_sum), (16, tax_sum), (17, total_sum)]:
        cell = ws.cell(row=total_row, column=col, value=val)
        cell.font = Font(name="Arial", bold=True, size=11, color=CLR_TOTAL_FG)
        cell.fill = _fill(CLR_TOTAL_BG)
        cell.border = _thin_border()
        cell.alignment = _right()
        cell.number_format = '#,##0.00'

    for col in range(1, len(SUMMARY_COLS) + 1):
        if col not in (1, 14, 16, 17):
            cell = ws.cell(row=total_row, column=col)
            if not cell.value:
                cell.fill = _fill(CLR_TOTAL_BG)
                cell.border = _thin_border()

    _set_row_height(ws, total_row, 22)

    # 凍結前3列
    ws.freeze_panes = f"A{DATA_START}"

    return total_sum


# ─────────────────────────────────────────
# 工作表 2：品項明細
# ─────────────────────────────────────────

DETAIL_COLS = [
    ("發票號碼",   "center", 16),
    ("日期",       "center", 13),
    ("廠商名稱",   "left",   22),
    ("費用科目",   "left",   14),
    ("部門",       "left",   10),
    ("品項說明",   "left",   35),
    ("數量",       "right",  8),
    ("單位",       "center", 8),
    ("單價",       "right",  12),
    ("小計",       "right",  12),
    ("幣別",       "center", 7),
]


def _build_detail_sheet(ws, invoices: list[dict]):
    ws.merge_cells("A1:K1")
    ws["A1"] = "品項明細"
    ws["A1"].font = Font(name="Arial", bold=True, size=14, color=CLR_HEADER_FG)
    ws["A1"].fill = _fill(CLR_HEADER_BG)
    ws["A1"].alignment = _center()
    _set_row_height(ws, 1, 30)

    HEADER_ROW = 2
    for col_idx, (label, align, width) in enumerate(DETAIL_COLS, start=1):
        cell = ws.cell(row=HEADER_ROW, column=col_idx, value=label)
        cell.font = Font(name="Arial", bold=True, color=CLR_HEADER_FG, size=10)
        cell.fill = _fill("2563EB")
        cell.alignment = _center()
        cell.border = _thin_border()
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    _set_row_height(ws, HEADER_ROW, 22)

    DATA_START = HEADER_ROW + 1
    r = DATA_START
    row_counter = 0

    for inv in invoices:
        items = inv.get("items", [])
        if not items:
            items = [{"description": "（無明細）", "quantity": "", "unit": "", "unit_price": "", "amount": ""}]

        for item in items:
            bg = CLR_ALT_ROW if row_counter % 2 == 0 else "FFFFFF"
            row_fill = _fill(bg)

            row_data = [
                inv.get("invoice_number", ""),
                inv.get("invoice_date", ""),
                inv.get("seller_name", ""),
                inv.get("account_category", ""),
                inv.get("department", ""),
                item.get("description", ""),
                item.get("quantity", ""),
                item.get("unit", ""),
                _safe_float(item.get("unit_price", 0)) or "",
                _safe_float(item.get("amount", 0)) or "",
                inv.get("currency", "TWD"),
            ]

            aligns = [a for _, a, _ in DETAIL_COLS]
            for col_idx, (val, align) in enumerate(zip(row_data, aligns), start=1):
                cell = ws.cell(row=r, column=col_idx, value=val)
                cell.font = _body_font()
                cell.fill = row_fill
                cell.border = _thin_border()
                if align == "center":
                    cell.alignment = _center()
                elif align == "right":
                    cell.alignment = _right()
                else:
                    cell.alignment = _left(wrap=True)

                if col_idx in (9, 10) and isinstance(val, float) and val:
                    cell.number_format = '#,##0.00'

            _set_row_height(ws, r, 18)
            r += 1
            row_counter += 1

    ws.freeze_panes = f"A{DATA_START}"


# ─────────────────────────────────────────
# 工作表 3：統計分析
# ─────────────────────────────────────────

def _build_stats_sheet(ws, invoices: list[dict]):
    ws.merge_cells("A1:H1")
    ws["A1"] = "統計分析"
    ws["A1"].font = Font(name="Arial", bold=True, size=14, color=CLR_HEADER_FG)
    ws["A1"].fill = _fill(CLR_HEADER_BG)
    ws["A1"].alignment = _center()
    _set_row_height(ws, 1, 30)

    current_row = 3

    def _section_title(row, text):
        ws.merge_cells(f"A{row}:H{row}")
        cell = ws[f"A{row}"]
        cell.value = text
        cell.font = Font(name="Arial", bold=True, size=11, color=CLR_SUBHDR_FG)
        cell.fill = _fill(CLR_SUBHDR_BG)
        cell.alignment = _left()
        cell.border = _thin_border()
        _set_row_height(ws, row, 20)
        return row + 1

    def _sub_header(row, labels, col_widths):
        for c, (label, width) in enumerate(zip(labels, col_widths), start=1):
            cell = ws.cell(row=row, column=c, value=label)
            cell.font = Font(name="Arial", bold=True, color=CLR_HEADER_FG, size=10)
            cell.fill = _fill("475569")
            cell.alignment = _center()
            cell.border = _thin_border()
            ws.column_dimensions[get_column_letter(c)].width = width
        _set_row_height(ws, row, 20)
        return row + 1

    def _data_row(row, values, bg="FFFFFF"):
        row_fill = _fill(bg)
        for c, val in enumerate(values, start=1):
            cell = ws.cell(row=row, column=c, value=val)
            cell.font = _body_font()
            cell.fill = row_fill
            cell.border = _thin_border()
            if c > 1 and isinstance(val, float):
                cell.alignment = _right()
                cell.number_format = '#,##0.00'
            else:
                cell.alignment = _left()
        _set_row_height(ws, row, 18)
        return row + 1

    # ── 1. 費用科目統計 ───────────────────────
    current_row = _section_title(current_row, "📊 費用科目統計")
    current_row = _sub_header(
        current_row,
        ["費用科目", "筆數", "未稅合計", "稅額合計", "含稅合計", "佔比(%)"],
        [18, 8, 15, 15, 15, 10],
    )

    acct_data: dict[str, dict] = {}
    grand_total = 0.0
    for inv in invoices:
        k = inv.get("account_category", "（未分類）") or "（未分類）"
        if k not in acct_data:
            acct_data[k] = {"count": 0, "subtotal": 0.0, "tax": 0.0, "total": 0.0}
        acct_data[k]["count"] += 1
        acct_data[k]["subtotal"] += _safe_float(inv.get("subtotal", 0))
        acct_data[k]["tax"] += _safe_float(inv.get("tax_amount", 0))
        acct_data[k]["total"] += _safe_float(inv.get("total_amount", 0))
        grand_total += _safe_float(inv.get("total_amount", 0))

    acct_start_row = current_row
    row_counter = 0
    for k, v in sorted(acct_data.items(), key=lambda x: -x[1]["total"]):
        ratio = (v["total"] / grand_total * 100) if grand_total else 0
        bg = CLR_ALT_ROW if row_counter % 2 == 0 else "FFFFFF"
        current_row = _data_row(current_row, [k, v["count"], v["subtotal"], v["tax"], v["total"], round(ratio, 1)], bg)
        row_counter += 1

    # 合計
    current_row = _data_row(
        current_row,
        ["合計", len(invoices),
         sum(v["subtotal"] for v in acct_data.values()),
         sum(v["tax"] for v in acct_data.values()),
         grand_total, 100.0],
        CLR_TOTAL_BG,
    )
    current_row += 1

    # ── 2. 部門統計 ───────────────────────────
    current_row = _section_title(current_row, "🏢 部門費用統計")
    current_row = _sub_header(
        current_row,
        ["部門", "筆數", "含稅合計", "佔比(%)"],
        [18, 8, 15, 10],
    )

    dept_data: dict[str, dict] = {}
    for inv in invoices:
        k = inv.get("department", "（未指定）") or "（未指定）"
        if k not in dept_data:
            dept_data[k] = {"count": 0, "total": 0.0}
        dept_data[k]["count"] += 1
        dept_data[k]["total"] += _safe_float(inv.get("total_amount", 0))

    row_counter = 0
    for k, v in sorted(dept_data.items(), key=lambda x: -x[1]["total"]):
        ratio = (v["total"] / grand_total * 100) if grand_total else 0
        bg = CLR_ALT_ROW if row_counter % 2 == 0 else "FFFFFF"
        current_row = _data_row(current_row, [k, v["count"], v["total"], round(ratio, 1)], bg)
        row_counter += 1
    current_row += 1

    # ── 3. 廠商統計 ───────────────────────────
    current_row = _section_title(current_row, "🏪 廠商消費統計 (前10名)")
    current_row = _sub_header(
        current_row,
        ["廠商名稱", "發票張數", "含稅合計", "平均金額"],
        [28, 10, 15, 15],
    )

    vendor_data: dict[str, dict] = {}
    for inv in invoices:
        k = inv.get("seller_name", "（未知廠商）") or "（未知廠商）"
        if k not in vendor_data:
            vendor_data[k] = {"count": 0, "total": 0.0}
        vendor_data[k]["count"] += 1
        vendor_data[k]["total"] += _safe_float(inv.get("total_amount", 0))

    row_counter = 0
    for k, v in sorted(vendor_data.items(), key=lambda x: -x[1]["total"])[:10]:
        avg = v["total"] / v["count"] if v["count"] else 0
        bg = CLR_ALT_ROW if row_counter % 2 == 0 else "FFFFFF"
        current_row = _data_row(current_row, [k, v["count"], v["total"], round(avg, 2)], bg)
        row_counter += 1
    current_row += 1

    # ── 4. 付款方式統計 ──────────────────────
    current_row = _section_title(current_row, "💳 付款方式統計")
    current_row = _sub_header(
        current_row,
        ["付款方式", "筆數", "金額合計", "佔比(%)"],
        [14, 8, 15, 10],
    )

    pay_data: dict[str, dict] = {}
    for inv in invoices:
        k = inv.get("payment_method", "（未填）") or "（未填）"
        if k not in pay_data:
            pay_data[k] = {"count": 0, "total": 0.0}
        pay_data[k]["count"] += 1
        pay_data[k]["total"] += _safe_float(inv.get("total_amount", 0))

    row_counter = 0
    for k, v in sorted(pay_data.items(), key=lambda x: -x[1]["total"]):
        ratio = (v["total"] / grand_total * 100) if grand_total else 0
        bg = CLR_ALT_ROW if row_counter % 2 == 0 else "FFFFFF"
        current_row = _data_row(current_row, [k, v["count"], v["total"], round(ratio, 1)], bg)
        row_counter += 1
    current_row += 1

    # ── 5. 總覽摘要 ──────────────────────────
    current_row = _section_title(current_row, "📋 總覽摘要")
    summary_items = [
        ("發票總張數", len(invoices)),
        ("含稅費用合計 (TWD)", f"NT$ {grand_total:,.0f}"),
        ("未稅費用合計", f"NT$ {sum(_safe_float(inv.get('subtotal',0)) for inv in invoices):,.0f}"),
        ("稅額合計", f"NT$ {sum(_safe_float(inv.get('tax_amount',0)) for inv in invoices):,.0f}"),
        ("費用科目數", len(acct_data)),
        ("部門數", len(dept_data)),
        ("廠商數", len(vendor_data)),
        ("製表時間", datetime.now().strftime("%Y/%m/%d %H:%M:%S")),
    ]

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 25
    row_counter = 0
    for label, val in summary_items:
        bg = CLR_ALT_ROW if row_counter % 2 == 0 else "FFFFFF"
        ws.cell(row=current_row, column=1, value=label).font = Font(name="Arial", bold=True, size=10)
        ws.cell(row=current_row, column=1).fill = _fill(bg)
        ws.cell(row=current_row, column=1).border = _thin_border()
        ws.cell(row=current_row, column=1).alignment = _left()
        ws.cell(row=current_row, column=2, value=val).font = Font(name="Arial", size=10)
        ws.cell(row=current_row, column=2).fill = _fill(bg)
        ws.cell(row=current_row, column=2).border = _thin_border()
        ws.cell(row=current_row, column=2).alignment = _left()
        _set_row_height(ws, current_row, 18)
        current_row += 1
        row_counter += 1


# ─────────────────────────────────────────
# 主要匯出函式
# ─────────────────────────────────────────

def create_invoice_excel(
    invoices: list[dict],
    report_title: str = "發票明細報表",
    company_name: str = "",
    period_start: str = "",
    period_end: str = "",
    preparer: str = "",
) -> bytes:
    """
    建立完整的發票 Excel 報表，回傳 bytes 供 Streamlit 下載。
    """
    wb = Workbook()

    # ── 工作表 1：彙總 ────────────────────────
    ws_summary = wb.active
    ws_summary.title = "發票彙總"
    ws_summary.sheet_view.showGridLines = True
    period_str = f"{period_start} ~ {period_end}" if (period_start or period_end) else ""
    _build_summary_sheet(
        ws_summary, invoices,
        title=report_title,
        company=company_name,
        period=period_str,
        preparer=preparer,
    )

    # ── 工作表 2：品項明細 ───────────────────
    ws_detail = wb.create_sheet("品項明細")
    _build_detail_sheet(ws_detail, invoices)

    # ── 工作表 3：統計分析 ───────────────────
    ws_stats = wb.create_sheet("統計分析")
    _build_stats_sheet(ws_stats, invoices)

    # ── 工作表分頁標籤顏色 ───────────────────
    ws_summary.sheet_properties.tabColor = "1E3A8A"
    ws_detail.sheet_properties.tabColor = "059669"
    ws_stats.sheet_properties.tabColor = "D97706"

    # ── 列印設定 ──────────────────────────────
    for ws in [ws_summary, ws_detail, ws_stats]:
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1
        ws.page_margins.left = 0.5
        ws.page_margins.right = 0.5
        ws.page_margins.top = 0.75
        ws.page_margins.bottom = 0.75

    # 輸出為 bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()
